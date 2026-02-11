#!/usr/bin/env python3
"""
Optimiert eine Besuchsroute fuer Adresslisten (CSV oder XLSX).

Standard:
- Startadresse: Vorderboden 1, 6373 Ennetbuergen
- Duplikate werden zusammengefasst (optional abschaltbar)
- Geocoding via Nominatim (OpenStreetMap) mit lokalem Cache
- Routenoptimierung: Nearest Neighbor + 2-opt
- Ausgabe im Format: StopNr, Adresse_raw, Strasse, PLZ, Ort, Hinweis
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple


DEFAULT_START = "..."
DEFAULT_USER_AGENT = "route-optimizer/1.0 (contact: local-script)"


@dataclass
class Stop:
    address_raw: str
    street: str
    plz: str
    city: str
    count: int
    lat: float
    lon: float


def normalize_whitespace(value: str) -> str:
    return " ".join((value or "").replace("\u00a0", " ").split()).strip()


def normalize_address_key(value: str) -> str:
    value = normalize_whitespace(value)
    value = value.replace(" ,", ",")
    return value.casefold()


def split_address(address: str) -> Tuple[str, str, str]:
    text = normalize_whitespace(address)
    match = re.match(r"^(.*?)(?:,)?\s+(\d{4})\s+(.+)$", text)
    if not match:
        return text, "", ""
    street = match.group(1).strip(" ,")
    plz = match.group(2).strip()
    city = match.group(3).strip()
    return street, plz, city


def detect_address_column(headers: Sequence[str], requested: Optional[str]) -> str:
    if not headers:
        raise ValueError("Die Datei hat keine Header-Zeile.")
    if requested:
        if requested in headers:
            return requested
        raise ValueError(f"Spalte '{requested}' nicht gefunden. Verfuegbare Spalten: {headers}")

    normalized: Dict[str, str] = {}
    for h in headers:
        normalized[normalize_whitespace(str(h)).casefold()] = h

    candidates = ("adress", "adresse", "address", "addr")
    for c in candidates:
        if c in normalized:
            return normalized[c]

    # Fallback: erste Spalte
    return headers[0]


def read_csv_rows(path: Path) -> List[Dict[str, str]]:
    last_error: Optional[Exception] = None
    encodings = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
    for enc in encodings:
        try:
            with path.open("r", encoding=enc, newline="") as fh:
                sample = fh.read(8192)
                fh.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=";,")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.DictReader(fh, dialect=dialect)
                rows = [dict(r) for r in reader]
                if reader.fieldnames:
                    return rows
        except Exception as exc:  # pragma: no cover - robust fallback
            last_error = exc
    if last_error:
        raise RuntimeError(f"CSV konnte nicht gelesen werden: {last_error}") from last_error
    raise RuntimeError("CSV konnte nicht gelesen werden.")


def read_xlsx_rows(path: Path) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError(
            "Fuer XLSX fehlt 'openpyxl'. Installieren mit: pip install openpyxl"
        ) from exc

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(iterator)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in raw_headers]
    rows: List[Dict[str, str]] = []
    for row in iterator:
        record: Dict[str, str] = {}
        for idx, value in enumerate(row):
            header = headers[idx] if idx < len(headers) else f"col_{idx}"
            record[header] = "" if value is None else str(value)
        rows.append(record)
    return rows


def read_input_rows(path: Path) -> List[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return read_xlsx_rows(path)
    raise ValueError(f"Dateityp nicht unterstuetzt: {path.suffix}. Bitte CSV oder XLSX verwenden.")


def extract_addresses(
    rows: Sequence[Dict[str, str]],
    requested_column: Optional[str],
) -> List[str]:
    if not rows:
        return []
    headers = list(rows[0].keys())
    col = detect_address_column(headers, requested_column)
    addresses: List[str] = []
    for row in rows:
        raw = row.get(col, "")
        addr = normalize_whitespace(raw)
        if addr:
            addresses.append(addr)
    return addresses


def load_cache(path: Path) -> Dict[str, Dict[str, float]]:
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}

    cache: Dict[str, Dict[str, float]] = {}
    if not isinstance(data, dict):
        return cache

    for key, value in data.items():
        if not isinstance(value, dict):
            continue
        lat = value.get("lat")
        lon = value.get("lon")
        if isinstance(lat, (int, float)) and isinstance(lon, (int, float)):
            cache[str(key)] = {"lat": float(lat), "lon": float(lon)}
    return cache


def save_cache(path: Path, cache: Dict[str, Dict[str, float]]) -> None:
    path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


class NominatimGeocoder:
    def __init__(self, user_agent: str, sleep_seconds: float = 1.1, countrycodes: str = "ch") -> None:
        self.user_agent = user_agent
        self.sleep_seconds = max(0.0, sleep_seconds)
        self.countrycodes = countrycodes
        self._last_call = 0.0

    def geocode(self, query: str, retries: int = 3) -> Optional[Tuple[float, float]]:
        for attempt in range(retries):
            wait = self.sleep_seconds - (time.time() - self._last_call)
            if wait > 0:
                time.sleep(wait)
            self._last_call = time.time()
            try:
                params = {
                    "q": query,
                    "format": "jsonv2",
                    "limit": 1,
                }
                if self.countrycodes:
                    params["countrycodes"] = self.countrycodes
                url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(params)
                req = urllib.request.Request(
                    url,
                    headers={
                        "User-Agent": self.user_agent,
                        "Accept": "application/json",
                    },
                )
                with urllib.request.urlopen(req, timeout=20) as response:
                    payload = response.read().decode("utf-8")
                data = json.loads(payload)
                if data:
                    lat = float(data[0]["lat"])
                    lon = float(data[0]["lon"])
                    return lat, lon
                return None
            except (urllib.error.URLError, TimeoutError, ValueError, KeyError, json.JSONDecodeError):
                if attempt == retries - 1:
                    return None
                time.sleep(1.5 * (attempt + 1))
        return None


def haversine_km(a: Tuple[float, float], b: Tuple[float, float]) -> float:
    lat1, lon1 = a
    lat2, lon2 = b
    r = 6371.0088
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    x = math.sin(dlat / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlon / 2) ** 2
    return 2 * r * math.asin(math.sqrt(x))


def path_length_km(order: Sequence[int], coords: Sequence[Tuple[float, float]], start: Tuple[float, float]) -> float:
    if not order:
        return 0.0
    total = haversine_km(start, coords[order[0]])
    for i in range(len(order) - 1):
        total += haversine_km(coords[order[i]], coords[order[i + 1]])
    return total


def nearest_neighbor_order(coords: Sequence[Tuple[float, float]], start: Tuple[float, float]) -> List[int]:
    remaining = set(range(len(coords)))
    order: List[int] = []
    current = start
    while remaining:
        nxt = min(remaining, key=lambda idx: haversine_km(current, coords[idx]))
        order.append(nxt)
        remaining.remove(nxt)
        current = coords[nxt]
    return order


def two_opt_open_path(
    order: List[int],
    coords: Sequence[Tuple[float, float]],
    start: Tuple[float, float],
    max_passes: int = 25,
) -> List[int]:
    n = len(order)
    if n < 4:
        return order

    for _ in range(max_passes):
        improved = False
        for i in range(0, n - 1):
            for j in range(i + 2, n + 1):
                prev_coord = start if i == 0 else coords[order[i - 1]]
                first_coord = coords[order[i]]
                last_coord = coords[order[j - 1]]
                next_coord = coords[order[j]] if j < n else None

                old_len = haversine_km(prev_coord, first_coord)
                new_len = haversine_km(prev_coord, last_coord)
                if next_coord is not None:
                    old_len += haversine_km(last_coord, next_coord)
                    new_len += haversine_km(first_coord, next_coord)

                if new_len + 1e-9 < old_len:
                    order[i:j] = reversed(order[i:j])
                    improved = True
        if not improved:
            break
    return order


def write_route_csv(
    output_path: Path,
    start_address: str,
    stops: Sequence[Stop],
) -> None:
    fieldnames = ["StopNr", "Adresse_raw", "Strasse", "PLZ", "Ort", "Hinweis"]
    with output_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        start_street, start_plz, start_city = split_address(start_address)
        writer.writerow(
            {
                "StopNr": 0,
                "Adresse_raw": start_address,
                "Strasse": start_street,
                "PLZ": start_plz,
                "Ort": start_city,
                "Hinweis": "START",
            }
        )
        for idx, stop in enumerate(stops, start=1):
            hint = f"{stop.count} Eintraege" if stop.count > 1 else ""
            writer.writerow(
                {
                    "StopNr": idx,
                    "Adresse_raw": stop.address_raw,
                    "Strasse": stop.street,
                    "PLZ": stop.plz,
                    "Ort": stop.city,
                    "Hinweis": hint,
                }
            )


def write_chunk_files(output_path: Path, start_address: str, stops: Sequence[Stop], chunk_size: int) -> List[Path]:
    if chunk_size <= 0:
        return []
    chunk_paths: List[Path] = []
    stem = output_path.stem
    for chunk_index, begin in enumerate(range(0, len(stops), chunk_size), start=1):
        end = begin + chunk_size
        chunk_stops = stops[begin:end]
        chunk_name = f"{stem}_chunk_{chunk_index}_max{chunk_size}.csv"
        chunk_path = output_path.with_name(chunk_name)
        write_route_csv(chunk_path, start_address, chunk_stops)
        chunk_paths.append(chunk_path)
    return chunk_paths


def write_unresolved(output_path: Path, unresolved: Sequence[str]) -> Optional[Path]:
    if not unresolved:
        return None
    path = output_path.with_name(f"{output_path.stem}_ungefunden.csv")
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Adresse_raw"])
        for addr in unresolved:
            writer.writerow([addr])
    return path


def aggregate_addresses(addresses: Sequence[str], keep_duplicates: bool, start_address: str) -> List[Tuple[str, int]]:
    start_key = normalize_address_key(start_address)
    if keep_duplicates:
        result: List[Tuple[str, int]] = []
        for addr in addresses:
            if normalize_address_key(addr) != start_key:
                result.append((addr, 1))
        return result

    ordered: Dict[str, Tuple[str, int]] = {}
    for addr in addresses:
        key = normalize_address_key(addr)
        if key == start_key:
            continue
        existing = ordered.get(key)
        if existing:
            ordered[key] = (existing[0], existing[1] + 1)
        else:
            ordered[key] = (addr, 1)
    return list(ordered.values())


def get_coordinates(
    address: str,
    cache: Dict[str, Dict[str, float]],
    geocoder: NominatimGeocoder,
) -> Optional[Tuple[float, float]]:
    key = normalize_address_key(address)
    cached = cache.get(key)
    if cached:
        return float(cached["lat"]), float(cached["lon"])

    coords = geocoder.geocode(address)
    if coords:
        cache[key] = {"lat": float(coords[0]), "lon": float(coords[1])}
    return coords


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Optimiert die effizienteste Route fuer eine Adressliste.")
    parser.add_argument("--input", required=True, help="Pfad zur Eingabedatei (CSV oder XLSX).")
    parser.add_argument(
        "--start",
        default=DEFAULT_START,
        help=f"Startadresse (Default: {DEFAULT_START})",
    )
    parser.add_argument("--output", default="route.csv", help="Ausgabe-CSV (Default: route.csv)")
    parser.add_argument(
        "--address-column",
        default=None,
        help="Name der Adress-Spalte. Wenn leer, wird automatisch erkannt.",
    )
    parser.add_argument("--cache", default="geocode_cache.json", help="Pfad zur Geocode-Cache-Datei.")
    parser.add_argument(
        "--keep-duplicates",
        action="store_true",
        help="Duplikat-Adressen nicht zusammenfassen.",
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=50,
        help="Optional: erzeugt zusaetzliche Chunk-Dateien mit max. N Stops (Default: 50, 0 deaktiviert).",
    )
    parser.add_argument(
        "--user-agent",
        default=DEFAULT_USER_AGENT,
        help="User-Agent fuer Nominatim API.",
    )
    parser.add_argument(
        "--countrycodes",
        default="ch",
        help="Laenderfilter fuer Geocoding (ISO-2, z.B. 'ch').",
    )
    parser.add_argument(
        "--sleep-seconds",
        type=float,
        default=1.1,
        help="Pause zwischen API-Aufrufen in Sekunden (Default: 1.1).",
    )
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    input_path = Path(args.input).expanduser()
    output_path = Path(args.output).expanduser()
    cache_path = Path(args.cache).expanduser()

    if not input_path.exists():
        print(f"Fehler: Eingabedatei nicht gefunden: {input_path}", file=sys.stderr)
        return 2

    try:
        rows = read_input_rows(input_path)
        addresses = extract_addresses(rows, args.address_column)
    except Exception as exc:
        print(f"Fehler beim Einlesen: {exc}", file=sys.stderr)
        return 2

    if not addresses:
        print("Keine Adressen gefunden.", file=sys.stderr)
        return 2

    aggregated = aggregate_addresses(addresses, args.keep_duplicates, args.start)
    print(f"Adressen geladen: {len(addresses)}")
    if not args.keep_duplicates:
        print(f"Einzigartige Stopps (ohne Start): {len(aggregated)}")

    cache = load_cache(cache_path)
    geocoder = NominatimGeocoder(
        user_agent=args.user_agent,
        sleep_seconds=args.sleep_seconds,
        countrycodes=args.countrycodes,
    )

    start_coords = get_coordinates(args.start, cache, geocoder)
    if not start_coords:
        print(f"Startadresse konnte nicht geocodiert werden: {args.start}", file=sys.stderr)
        return 3

    valid_stops: List[Stop] = []
    unresolved: List[str] = []
    total = len(aggregated)

    for idx, (addr, count) in enumerate(aggregated, start=1):
        coords = get_coordinates(addr, cache, geocoder)
        if not coords:
            unresolved.append(addr)
            print(f"[{idx}/{total}] Nicht gefunden: {addr}")
            continue
        street, plz, city = split_address(addr)
        valid_stops.append(
            Stop(
                address_raw=addr,
                street=street,
                plz=plz,
                city=city,
                count=count,
                lat=coords[0],
                lon=coords[1],
            )
        )
        print(f"[{idx}/{total}] OK: {addr}")

    save_cache(cache_path, cache)

    if not valid_stops:
        print("Keine gueltigen Stopps geocodiert. Abbruch.", file=sys.stderr)
        unresolved_file = write_unresolved(output_path, unresolved)
        if unresolved_file:
            print(f"Ungefundene Adressen: {unresolved_file}")
        return 4

    coords = [(s.lat, s.lon) for s in valid_stops]
    initial_order = list(range(len(valid_stops)))
    nn_order = nearest_neighbor_order(coords, start_coords)
    best_order = two_opt_open_path(nn_order, coords, start_coords)

    before_km = path_length_km(initial_order, coords, start_coords)
    after_km = path_length_km(best_order, coords, start_coords)

    ordered_stops = [valid_stops[i] for i in best_order]
    write_route_csv(output_path, args.start, ordered_stops)
    chunk_paths = write_chunk_files(output_path, args.start, ordered_stops, args.chunk_size)
    unresolved_file = write_unresolved(output_path, unresolved)

    print("")
    print(f"Hauptdatei geschrieben: {output_path}")
    if chunk_paths:
        print(f"Chunk-Dateien: {len(chunk_paths)}")
        for p in chunk_paths:
            print(f" - {p}")
    if unresolved_file:
        print(f"Ungefundene Adressen: {unresolved_file} ({len(unresolved)})")
    print(f"Stops in Route: {len(ordered_stops)}")
    print(f"Distanz vor Optimierung: {before_km:.2f} km")
    print(f"Distanz nach Optimierung: {after_km:.2f} km")
    if before_km > 0:
        print(f"Verbesserung: {(before_km - after_km):.2f} km ({(1 - after_km / before_km) * 100:.1f}%)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
